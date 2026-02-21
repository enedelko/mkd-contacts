"""
CORE-01: Импорт реестра помещений и контактов (CSV/XLS/XLSX).
Парсинг, валидация, сопоставление по Blind Index, шифрование (BE-02), отчёт.
"""
import csv
import io
import logging
import random
import re
from typing import Any

from sqlalchemy import text

from app.crypto import (
    blind_index_email,
    blind_index_phone,
    blind_index_telegram_id,
    decrypt,
    encrypt,
)
from app.db import get_db
from app.room_normalizer import normalize_room_number

logger = logging.getLogger(__name__)

# Ожидаемые колонки (LOST-02, CORE-01). Минимальный набор для валидации структуры.
REQUIRED_PREMISE_COLUMNS = ["cadastral_number"]
OPTIONAL_PREMISE_COLUMNS = ["area", "entrance", "floor", "premises_type", "premises_number"]
CONTACT_COLUMNS = ["phone", "email", "telegram_id", "how_to_address"]
# ADM-06: опциональные колонки загрузки контактов
CONTACTS_ONLY_OPTIONAL = ["is_owner", "barrier_vote", "vote_format", "registered_in_ed"]
# Синонимы для маппинга заголовков файла
COLUMN_ALIASES = {
    "cadastral_number": ["cadastral_number", "cadastral", "кадастровый_номер", "кадастровый номер"],
    "area": ["area", "площадь"],
    "entrance": ["entrance", "подъезд", "entrance_number"],
    "floor": ["floor", "этаж"],
    "premises_type": ["premises_type", "premise_type", "тип_помещения", "тип помещения"],
    "premises_number": ["premises_number", "premise_number", "room_number", "номер_помещения", "номер помещения", "кв", "квартира"],
    "phone": ["phone", "телефон", "tel"],
    "email": ["email", "почта", "e-mail"],
    "telegram_id": ["telegram_id", "telegram", "тг"],
    "how_to_address": ["how_to_address", "как_обращаться", "обращение", "как обращаться"],
    "is_owner": ["is_owner", "собственник", "Собственник?", "проживающий"],
    "barrier_vote": ["barrier_vote", "позиция_по_шлагбаумам", "позиция по шлагбаумам"],
    "vote_format": ["vote_format", "формат_голосования", "формат голосования"],
    "registered_in_ed": ["registered_in_ed", "зарегистрирован_в_эд", "зарегистрирован в эд", "электронный_дом", "электронный дом", "эл_дом"],
}


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_").replace("-", "_")


def _map_headers(headers: list[str]) -> dict[str, int]:
    """Сопоставить заголовки файла с каноническими именами колонок."""
    canonical_to_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        norm = _normalize_header(h)
        for canonical, aliases in COLUMN_ALIASES.items():
            if norm in [ _normalize_header(a) for a in aliases ] or norm == canonical:
                if canonical not in canonical_to_idx:
                    canonical_to_idx[canonical] = i
                break
    return canonical_to_idx


def get_expected_columns() -> list[str]:
    """Список ожидаемых колонок для ответа API (LOST-02)."""
    return REQUIRED_PREMISE_COLUMNS + OPTIONAL_PREMISE_COLUMNS + CONTACT_COLUMNS


def get_expected_columns_contacts_only() -> list[str]:
    """Колонки для загрузки только контактов (ADM-06): кадастр + хотя бы один контактный столбец."""
    return ["cadastral_number", "phone", "email", "telegram_id", "how_to_address"] + CONTACTS_ONLY_OPTIONAL


def _normalize_bool(val: Any) -> bool | None:
    """Привести значение из файла к bool для is_owner. None если пусто или не распознано."""
    if val is None:
        return None
    s = (str(val).strip().lower() if hasattr(val, "strip") else str(val).strip().lower()) if val else ""
    if not s:
        return None
    if s in ("1", "true", "yes", "да", "да."):
        return True
    if s in ("0", "false", "no", "нет", "нет."):
        return False
    return None


def _normalize_barrier_vote(val: Any) -> str | None:
    """Привести значение к for | against | undecided. None если пусто."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    if s in ("for", "за", "да"):
        return "for"
    if s in ("against", "против", "нет"):
        return "against"
    if s in ("undecided", "не определён", "не определен", "не определился", "воздержался", "воздержалась", "воздержались",
             "пока не решил", "пока не решила", "пока не решили", "пока не определились", "думают"):
        return "undecided"
    return None


def _normalize_vote_format(val: Any) -> str | None:
    """Привести значение к electronic | paper | undecided. None если пусто."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    if s in ("electronic", "электронно", "электронный"):
        return "electronic"
    if s in ("paper", "бумага", "бумажный", "на бумаге"):
        return "paper"
    if s in ("undecided", "не определён", "не определен", "не определился", "пока не решили", "пока не определились", "думают"):
        return "undecided"
    return None


def _normalize_registered_ed(val: Any) -> str | None:
    """Привести значение к yes | no. None если пусто или не распознано."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    if s in ("yes", "да", "true", "1"):
        return "yes"
    if s in ("no", "нет", "false", "0"):
        return "no"
    return None


def validate_structure(detected: list[str]) -> tuple[bool, list[str], list[str]]:
    """
    Проверка структуры: обязательна хотя бы cadastral_number.
    Возвращает (ok, expected, detected).
    """
    expected = get_expected_columns()
    detected_norm = [_normalize_header(c) for c in detected]
    expected_norm = [_normalize_header(c) for c in expected]
    has_cadastral = any(
        _normalize_header(a) in detected_norm
        for aliases in [COLUMN_ALIASES["cadastral_number"]]
        for a in aliases
    )
    if not has_cadastral:
        return False, expected, detected
    return True, expected, detected


def _row_to_dict(row: list[Any], mapping: dict[str, int]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, idx in mapping.items():
        if idx < len(row):
            val = row[idx]
            if hasattr(val, "strip"):
                val = val.strip() if val else ""
            elif val is not None:
                val = str(val).strip()
            else:
                val = ""
            out[key] = val or None
    return out


def _read_csv(content: bytes) -> tuple[list[str], list[list[Any]]]:
    """Прочитать CSV UTF-8 (SR-CORE01-002)."""
    text_content = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text_content), delimiter=";")
    rows = list(reader)
    if not rows:
        raise ValueError("Empty file")
    headers = [h.strip() for h in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def _read_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    """Прочитать первый лист XLSX (SR-CORE01-003)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        raise ValueError("No sheet")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty sheet")
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data_rows = [[str(c).strip() if c is not None else "" for c in r] for r in rows[1:]]
    return headers, data_rows


def _read_xls(content: bytes) -> tuple[list[str], list[list[Any]]]:
    """Прочитать первый лист XLS (SR-CORE01-003)."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=content)
    sheet = wb.sheet_by_index(0)
    headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    data_rows = [
        [str(sheet.cell_value(r, c)).strip() if sheet.cell_value(r, c) else "" for c in range(sheet.ncols)]
        for r in range(1, sheet.nrows)
    ]
    return headers, data_rows


def parse_file(content: bytes, filename: str) -> tuple[list[str], list[str], list[dict[str, Any]]]:
    """
    Определить формат по расширению/содержимому, распарсить.
    Возвращает (original_headers, canonical_columns, list of row dicts).
    """
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        headers, data_rows = _read_csv(content)
    elif fn.endswith(".xlsx"):
        headers, data_rows = _read_xlsx(content)
    elif fn.endswith(".xls"):
        headers, data_rows = _read_xls(content)
    else:
        if content[:4] == b"PK\x03\x04":
            headers, data_rows = _read_xlsx(content)
        else:
            headers, data_rows = _read_csv(content)
    original_headers = [h for h in headers if (h or "").strip()]
    mapping = _map_headers(headers)
    rows = [_row_to_dict(r, mapping) for r in data_rows]
    return original_headers, list(mapping.keys()), rows


def _get_or_create_premise(
    db, cadastral_number: str, area, entrance, floor, premises_type, premises_number: str
) -> str:
    """Получить кадастровый номер помещения (существующего или созданного) (SR-CORE01-006)."""
    # Проверка существования строки: не вставлять дубликат по PK при повторном появлении кадастра в импорте.
    row = db.execute(
        text("SELECT 1 FROM premises WHERE cadastral_number = :cn"),
        {"cn": cadastral_number},
    ).fetchone()
    if row:
        return cadastral_number
    db.execute(
        text(
            "INSERT INTO premises (cadastral_number, area, entrance, floor, premises_type, premises_number) "
            "VALUES (:cn, :area, :entrance, :floor, :pt, :pn)"
        ),
        {
            "cn": cadastral_number,
            "area": area,
            "entrance": entrance,
            "floor": floor,
            "pt": premises_type,
            "pn": premises_number,
        },
    )
    db.flush()
    return cadastral_number


def _find_contact_by_indexes(db, premise_id: str, phone_idx, email_idx, telegram_id_idx) -> dict | None:
    """Найти контакт по premise_id (cadastral_number) и любому из Blind Index. Возвращает id, индексы (для коллизии) и флаги заполненности (для обогащения SR-CORE01-014)."""
    cols = "id, phone_idx, email_idx, telegram_id_idx, (COALESCE(trim(phone),'') != '') as has_phone, (COALESCE(trim(email),'') != '') as has_email, (COALESCE(trim(telegram_id),'') != '') as has_telegram_id, (COALESCE(trim(how_to_address),'') != '') as has_how"
    if phone_idx:
        r = db.execute(
            text(f"SELECT {cols} FROM contacts WHERE premise_id = :pid AND phone_idx = :pi"),
            {"pid": premise_id, "pi": phone_idx},
        ).fetchone()
        if r:
            return {"id": r[0], "phone_idx": r[1], "email_idx": r[2], "telegram_id_idx": r[3], "has_phone": r[4], "has_email": r[5], "has_telegram_id": r[6], "has_how": r[7]}
    if email_idx:
        r = db.execute(
            text(f"SELECT {cols} FROM contacts WHERE premise_id = :pid AND email_idx = :ei"),
            {"pid": premise_id, "ei": email_idx},
        ).fetchone()
        if r:
            return {"id": r[0], "phone_idx": r[1], "email_idx": r[2], "telegram_id_idx": r[3], "has_phone": r[4], "has_email": r[5], "has_telegram_id": r[6], "has_how": r[7]}
    if telegram_id_idx:
        r = db.execute(
            text(f"SELECT {cols} FROM contacts WHERE premise_id = :pid AND telegram_id_idx = :ti"),
            {"pid": premise_id, "ti": telegram_id_idx},
        ).fetchone()
        if r:
            return {"id": r[0], "phone_idx": r[1], "email_idx": r[2], "telegram_id_idx": r[3], "has_phone": r[4], "has_email": r[5], "has_telegram_id": r[6], "has_how": r[7]}
    return None


def _collision(existing: dict, row: dict, phone_idx, email_idx, telegram_id_idx) -> str | None:
    """SR-CORE01-016: тот же индекс, но в файле другой идентификатор противоречит БД (уже заполнен иным значением)."""
    if not existing:
        return None
    reasons = []
    if row.get("phone") and phone_idx and existing.get("email_idx") and email_idx and existing.get("email_idx") != email_idx:
        reasons.append("phone matches existing contact, email contradicts")
    if row.get("email") and email_idx and existing.get("phone_idx") and phone_idx and existing.get("phone_idx") != phone_idx:
        reasons.append("email matches existing contact, phone contradicts")
    if row.get("telegram_id") and telegram_id_idx:
        if existing.get("phone_idx") and phone_idx and existing.get("phone_idx") != phone_idx:
            reasons.append("telegram_id matches existing contact, phone contradicts")
        if existing.get("email_idx") and email_idx and existing.get("email_idx") != email_idx:
            reasons.append("telegram_id matches existing contact, email contradicts")
    return "; ".join(reasons) if reasons else None


def run_import(rows: list[dict[str, Any]], client_ip: str | None = None) -> dict[str, Any]:
    """
    Выполнить импорт в транзакции. Все валидные строки записываются; ошибки по строкам в отчёте.
    Возвращает { accepted, rejected, errors: [ { row, message } ] }.
    """
    accepted = 0
    rejected = 0
    errors: list[dict[str, Any]] = []
    with get_db() as db:
        try:
            for row_num, row in enumerate(rows, start=2):
                row_1based = row_num
                cadastral = (row.get("cadastral_number") or "").strip()
                if not cadastral:
                    errors.append({"row": row_1based, "message": "Missing required field: cadastral_number"})
                    rejected += 1
                    continue
                phone = (row.get("phone") or "").strip() or None
                email = (row.get("email") or "").strip() or None
                telegram_id = (row.get("telegram_id") or "").strip() or None
                how_to_address = (row.get("how_to_address") or "").strip() or None
                has_contact = phone or email or telegram_id
                premises_number_raw = (row.get("premises_number") or "").strip() or None
                premises_number = normalize_room_number(premises_number_raw) or premises_number_raw or ""
                try:
                    premise_id = _get_or_create_premise(
                        db,
                        cadastral,
                        row.get("area") or None,
                        row.get("entrance") or None,
                        row.get("floor") or None,
                        row.get("premises_type") or None,
                        premises_number,
                    )
                except Exception as e:
                    errors.append({"row": row_1based, "message": f"Premise error: {e}"})
                    rejected += 1
                    continue
                # Если контактных данных нет — помещение создано, контакт не добавляется
                if not has_contact:
                    accepted += 1
                    continue
                phone_idx = blind_index_phone(phone) if phone else None
                email_idx = blind_index_email(email) if email else None
                telegram_id_idx = blind_index_telegram_id(telegram_id) if telegram_id else None
                existing = _find_contact_by_indexes(db, premise_id, phone_idx, email_idx, telegram_id_idx)
                collision_msg = _collision(existing, row, phone_idx, email_idx, telegram_id_idx)
                if collision_msg:
                    errors.append({"row": row_1based, "message": f"Collision: {collision_msg}"})
                    rejected += 1
                    continue
                phone_enc = encrypt(phone)
                email_enc = encrypt(email)
                telegram_id_enc = encrypt(telegram_id)
                how_enc = encrypt(how_to_address)
                if existing:
                    contact_id = existing["id"]
                    updates = []
                    params = {"cid": contact_id}
                    if email_enc and not existing.get("has_email"):
                        updates.append("email = :email"); params["email"] = email_enc; params["email_idx"] = email_idx
                    if phone_enc and not existing.get("has_phone"):
                        updates.append("phone = :phone"); params["phone"] = phone_enc; params["phone_idx"] = phone_idx
                    if telegram_id_enc and not existing.get("has_telegram_id"):
                        updates.append("telegram_id = :telegram_id"); params["telegram_id"] = telegram_id_enc; params["telegram_id_idx"] = telegram_id_idx
                    if how_enc and not existing.get("has_how"):
                        updates.append("how_to_address = :how_to_address"); params["how_to_address"] = how_enc
                    if updates:
                        set_parts = updates.copy()
                        if "email_idx" in params: set_parts.append("email_idx = :email_idx")
                        if "phone_idx" in params: set_parts.append("phone_idx = :phone_idx")
                        if "telegram_id_idx" in params: set_parts.append("telegram_id_idx = :telegram_id_idx")
                        set_parts.append("updated_at = CURRENT_TIMESTAMP")
                        db.execute(text("UPDATE contacts SET " + ", ".join(set_parts) + " WHERE id = :cid"), params)
                    accepted += 1
                else:
                    db.execute(
                        text(
                            "INSERT INTO contacts (premise_id, is_owner, phone, email, telegram_id, how_to_address, "
                            "phone_idx, email_idx, telegram_id_idx, status, ip) "
                            "VALUES (:pid, true, :phone, :email, :telegram_id, :how_to_address, "
                            ":phone_idx, :email_idx, :telegram_id_idx, 'pending', :ip)"
                        ),
                        {
                            "pid": premise_id,
                            "phone": phone_enc,
                            "email": email_enc,
                            "telegram_id": telegram_id_enc,
                            "how_to_address": how_enc,
                            "phone_idx": phone_idx,
                            "email_idx": email_idx,
                            "telegram_id_idx": telegram_id_idx,
                            "ip": client_ip,
                        },
                    )
                    accepted += 1
            db.commit()
        except Exception as e:
            db.rollback()
            logger.exception("Import transaction failed")
            raise
    return {"accepted": accepted, "rejected": rejected, "errors": errors}


def run_import_contacts_only(rows: list[dict[str, Any]], client_ip: str | None = None) -> dict[str, Any]:
    """
    ADM-06: Импорт только контактов. Помещения не создаются и не обновляются.
    Обязательны cadastral_number и хотя бы одно из: phone, email, telegram_id.
    Опционально: is_owner, barrier_vote, vote_format.
    """
    accepted = 0
    rejected = 0
    errors: list[dict[str, Any]] = []
    with get_db() as db:
        try:
            for row_num, row in enumerate(rows, start=2):
                row_1based = row_num
                cadastral = (row.get("cadastral_number") or "").strip()
                if not cadastral:
                    errors.append({"row": row_1based, "message": "Missing required field: cadastral_number"})
                    rejected += 1
                    continue
                phone = (row.get("phone") or "").strip() or None
                email = (row.get("email") or "").strip() or None
                telegram_id = (row.get("telegram_id") or "").strip() or None
                how_to_address = (row.get("how_to_address") or "").strip() or None
                has_contact = phone or email or telegram_id
                if not has_contact:
                    errors.append({"row": row_1based, "message": "At least one of phone, email, telegram_id is required"})
                    rejected += 1
                    continue
                premise_row = db.execute(
                    text("SELECT 1 FROM premises WHERE cadastral_number = :cn"),
                    {"cn": cadastral},
                ).fetchone()
                if not premise_row:
                    errors.append({"row": row_1based, "message": f"Premise not found: {cadastral}"})
                    rejected += 1
                    continue
                premise_id = cadastral
                is_owner_val = _normalize_bool(row.get("is_owner"))
                if is_owner_val is None:
                    is_owner_val = True
                barrier_vote_val = _normalize_barrier_vote(row.get("barrier_vote"))
                vote_format_val = _normalize_vote_format(row.get("vote_format"))
                registered_ed_val = _normalize_registered_ed(row.get("registered_in_ed"))

                phone_idx = blind_index_phone(phone) if phone else None
                email_idx = blind_index_email(email) if email else None
                telegram_id_idx = blind_index_telegram_id(telegram_id) if telegram_id else None
                existing = _find_contact_by_indexes(db, premise_id, phone_idx, email_idx, telegram_id_idx)
                collision_msg = _collision(existing, row, phone_idx, email_idx, telegram_id_idx)
                if collision_msg:
                    errors.append({"row": row_1based, "message": f"Collision: {collision_msg}"})
                    rejected += 1
                    continue
                phone_enc = encrypt(phone)
                email_enc = encrypt(email)
                telegram_id_enc = encrypt(telegram_id)
                how_enc = encrypt(how_to_address)
                if existing:
                    contact_id = existing["id"]
                    updates = []
                    params = {"cid": contact_id, "io": is_owner_val}
                    if email_enc and not existing.get("has_email"):
                        updates.append("email = :email"); params["email"] = email_enc; params["email_idx"] = email_idx
                    if phone_enc and not existing.get("has_phone"):
                        updates.append("phone = :phone"); params["phone"] = phone_enc; params["phone_idx"] = phone_idx
                    if telegram_id_enc and not existing.get("has_telegram_id"):
                        updates.append("telegram_id = :telegram_id"); params["telegram_id"] = telegram_id_enc; params["telegram_id_idx"] = telegram_id_idx
                    if how_enc and not existing.get("has_how"):
                        updates.append("how_to_address = :how_to_address"); params["how_to_address"] = how_enc
                    updates.append("is_owner = :io")
                    if registered_ed_val is not None:
                        updates.append("registered_in_ed = :re"); params["re"] = registered_ed_val
                    if updates:
                        set_parts = updates.copy()
                        if "email_idx" in params: set_parts.append("email_idx = :email_idx")
                        if "phone_idx" in params: set_parts.append("phone_idx = :phone_idx")
                        if "telegram_id_idx" in params: set_parts.append("telegram_id_idx = :telegram_id_idx")
                        set_parts.append("updated_at = CURRENT_TIMESTAMP")
                        db.execute(text("UPDATE contacts SET " + ", ".join(set_parts) + " WHERE id = :cid"), params)
                    _upsert_oss_voting(db, contact_id, barrier_vote_val, vote_format_val)
                    accepted += 1
                else:
                    db.execute(
                        text(
                            "INSERT INTO contacts (premise_id, is_owner, phone, email, telegram_id, how_to_address, "
                            "phone_idx, email_idx, telegram_id_idx, registered_in_ed, status, ip) "
                            "VALUES (:pid, :io, :phone, :email, :telegram_id, :how_to_address, "
                            ":phone_idx, :email_idx, :telegram_id_idx, :re, 'pending', :ip)"
                        ),
                        {
                            "pid": premise_id,
                            "io": is_owner_val,
                            "phone": phone_enc,
                            "email": email_enc,
                            "telegram_id": telegram_id_enc,
                            "how_to_address": how_enc,
                            "phone_idx": phone_idx,
                            "email_idx": email_idx,
                            "telegram_id_idx": telegram_id_idx,
                            "re": registered_ed_val,
                            "ip": client_ip,
                        },
                    )
                    db.flush()
                    contact_id = db.execute(text("SELECT id FROM contacts WHERE premise_id = :pid ORDER BY id DESC LIMIT 1"), {"pid": premise_id}).scalar()
                    if contact_id:
                        _upsert_oss_voting(db, contact_id, barrier_vote_val, vote_format_val)
                    accepted += 1
            db.commit()
        except Exception as e:
            db.rollback()
            logger.exception("Contacts-only import transaction failed")
            raise
    return {"accepted": accepted, "rejected": rejected, "errors": errors}


def _upsert_oss_voting(db, contact_id: int, barrier_vote: str | None, vote_format: str | None) -> None:
    """Вставить или обновить запись oss_voting для контакта. None не затирает существующее."""
    if barrier_vote is None and vote_format is None:
        return
    row = db.execute(text("SELECT barrier_vote, vote_format FROM oss_voting WHERE contact_id = :cid"), {"cid": contact_id}).fetchone()
    bv = barrier_vote or (row[0] if row else "undecided")
    vf = vote_format or (row[1] if row else "undecided")
    if row:
        db.execute(
            text("UPDATE oss_voting SET barrier_vote = :bv, vote_format = :vf WHERE contact_id = :cid"),
            {"bv": bv, "vf": vf, "cid": contact_id},
        )
    else:
        db.execute(
            text("INSERT INTO oss_voting (contact_id, barrier_vote, vote_format, voted) VALUES (:cid, :bv, :vf, false)"),
            {"cid": contact_id, "bv": bv, "vf": vf},
        )


# ADM-08: колонки шаблона контактов (совместимы с ADM-06). + колонка «ТГ» (ссылка, SR-ADM08-006).
CONTACTS_TEMPLATE_HEADERS = [
    "cadastral_number", "premises_type", "premises_number",
    "phone", "email", "telegram_id", "how_to_address",
    "is_owner", "barrier_vote", "vote_format", "registered_in_ed",
]
CONTACTS_TEMPLATE_HEADERS_RU = [
    "кадастровый номер", "тип помещения", "номер помещения",
    "телефон", "почта", "telegram_id", "Ссылка ТГ", "обращение",
    "Собственник?", "позиция по шлагбаумам", "формат голосования", "Электронный дом",
]
COL_TG_LINK = 6  # индекс колонки «Ссылка ТГ» (0-based); при импорте не используется (SR-ADM08-003)
COL_PHONE = 3   # колонка телефон (D)
COL_TELEGRAM_ID = 5  # колонка telegram_id (F)
COL_BARRIER_VOTE = 9   # колонка позиция по шлагбаумам (J)
COL_VOTE_FORMAT = 10   # колонка формат голосования (K)
COL_REGISTERED_ED = 11  # колонка Электронный дом (L)

# Транслитерация кириллицы в латиницу для имени файла (ГОСТ 7.79-2000, тип B)
TRANSLIT_RU_EN = str.maketrans({
    "А": "A", "Б": "B", "В": "V", "Г": "G", "Д": "D", "Е": "E", "Ё": "E", "Ж": "Zh", "З": "Z",
    "И": "I", "Й": "J", "К": "K", "Л": "L", "М": "M", "Н": "N", "О": "O", "П": "P", "Р": "R",
    "С": "S", "Т": "T", "У": "U", "Ф": "F", "Х": "Kh", "Ц": "Ts", "Ч": "Ch", "Ш": "Sh", "Щ": "Shch",
    "Ъ": "", "Ы": "Y", "Ь": "", "Э": "E", "Ю": "Yu", "Я": "Ya",
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh", "з": "z",
    "и": "i", "й": "j", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
    "с": "s", "т": "t", "у": "u", "ф": "f", "х": "kh", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "shch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
})

# Русские подписи для выгрузки шаблона (как в списке контактов CORE-03)
BARRIER_VOTE_DISPLAY = {"for": "ЗА", "against": "Против", "undecided": "Не определился"}
VOTE_FORMAT_DISPLAY = {"electronic": "Электронно", "paper": "Бумага", "undecided": "Не определился"}
REGISTERED_ED_DISPLAY = {"yes": "Да", "no": "Нет"}

# Canary: имена-гомографы (латиница/кириллица), выглядят как обычные имена
CANARY_NAMES = [
    "Aлeкceй",   # Алексей
    "Дмитpий",   # Дмитрий
    "Cеpгей",    # Сергей
    "Андpей",    # Андрей
    "Михaил",    # Михаил
    "Никоlай",   # Николай
    "Aлeкcандp", # Александр
    "Евгeний",   # Евгений
    "Ивaн",      # Иван
    "Пaвел",     # Павел
]


def create_watermark(admin_telegram_id: str, entrance: str) -> dict[str, Any] | None:
    """
    Создать запись в export_watermarks: случайное помещение подъезда, телефон +7 916 XXXXXXX,
    canary_telegram_id, одно из CANARY_NAMES. Возвращает dict с ключами premise_id, phone,
    canary_telegram_id, how_to_address для вставки в шаблон и список контактов.
    """
    with get_db() as db:
        premise_row = db.execute(
            text(
                "SELECT cadastral_number FROM premises WHERE entrance = :e ORDER BY random() LIMIT 1"
            ),
            {"e": entrance.strip()},
        ).fetchone()
        if not premise_row:
            return None
        premise_id = premise_row[0]
        phone = "+7916" + "".join(random.choices("0123456789", k=7))
        canary_telegram_id = str(random.randint(10**7, 10**10 - 1))
        how_to_address = random.choice(CANARY_NAMES)
        result = db.execute(
            text(
                "INSERT INTO export_watermarks "
                "(admin_telegram_id, entrance, premise_id, phone, canary_telegram_id, how_to_address) "
                "VALUES (:aid, :e, :pid, :phone, :tg, :how) "
                "RETURNING created_at"
            ),
            {
                "aid": admin_telegram_id,
                "e": entrance.strip(),
                "pid": premise_id,
                "phone": phone,
                "tg": canary_telegram_id,
                "how": how_to_address,
            },
        )
        created_at_row = result.fetchone()
        db.commit()
    created_at = created_at_row[0] if created_at_row else None
    return {
        "premise_id": premise_id,
        "phone": phone,
        "canary_telegram_id": canary_telegram_id,
        "how_to_address": how_to_address,
        "created_at": created_at.isoformat() if created_at else None,
    }


def _telegram_link(telegram_id: str | None, phone: str | None) -> str:
    """SR-ADM08-006: ссылка в Telegram по telegram_id или по телефону (логика как в списке контактов)."""
    if telegram_id and str(telegram_id).strip():
        tid = str(telegram_id).strip()
        if tid.isdigit():
            return f"tg://user?id={tid}"
        return f"https://t.me/{tid.lstrip('@')}"
    if phone and str(phone).strip():
        digits = "".join(c for c in str(phone) if c.isdigit())
        if len(digits) >= 10:
            if digits.startswith("8") and len(digits) == 11:
                digits = "7" + digits[1:]
            if digits.startswith("7") and len(digits) == 11:
                return f"https://t.me/+{digits}"
            return f"https://t.me/+{digits}"
    return ""


def transliterate_entrance_for_filename(entrance: str) -> str:
    """Транслитерация подъезда для имени файла: кириллица → латиница (ГОСТ 7.79-2000), остальное — буквы/цифры/дефис."""
    if not entrance:
        return "entrance"
    s = str(entrance).strip().translate(TRANSLIT_RU_EN)
    s = re.sub(r"[^\w\-.]", "_", s)
    return s.strip("_") or "entrance"


def _format_phone_display(phone: str | None) -> str:
    """Формат +7 (XXX) XXX-XX-XX для отображения в XLSX (опционально, п. 8)."""
    if not phone or not str(phone).strip():
        return ""
    digits = "".join(c for c in str(phone) if c.isdigit())
    if len(digits) == 11 and digits.startswith("7"):
        return f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
        return f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    return phone


def build_contacts_template_xlsx(entrance: str, canary_row: dict[str, Any] | None = None) -> tuple[bytes, int]:
    """
    ADM-08: Сформировать XLSX-шаблон контактов по подъезду.
    Одна строка на контакт; при отсутствии контакта — одна строка с пустыми полями контакта.
    canary_row: опционально dict с premise_id, phone, canary_telegram_id, how_to_address — вставляется как строка по этому помещению.
    Возвращает (содержимое файла, количество строк данных).
    """
    import openpyxl

    rows: list[list[Any]] = []
    with get_db() as db:
        # SR-ADM08-005: сортировка как в CORE-03 — по типу помещения, затем по номеру (естественная)
        q = text(
            "SELECT p.cadastral_number, p.premises_type, p.premises_number, "
            "c.id, c.phone, c.email, c.telegram_id, c.how_to_address, c.is_owner, "
            "o.barrier_vote, o.vote_format, c.registered_in_ed "
            "FROM premises p "
            "LEFT JOIN contacts c ON c.premise_id = p.cadastral_number "
            "LEFT JOIN oss_voting o ON o.contact_id = c.id "
            "WHERE p.entrance = :e "
            "ORDER BY p.premises_type NULLS LAST, "
            "(NULLIF(TRIM(REGEXP_REPLACE(COALESCE(p.premises_number, ''), '[^0-9].*', '')), '')::int) NULLS LAST, "
            "p.premises_number NULLS LAST, c.id NULLS LAST"
        )
        result = db.execute(q, {"e": entrance}).fetchall()
        for r in result:
            cn, pt, pn = r[0], r[1], r[2]
            cid, phone_enc, email_enc, tg_enc, how_enc = r[3], r[4], r[5], r[6], r[7]
            is_owner, bv, vf, reg_ed = r[8], r[9], r[10], r[11]
            if cid is None:
                rows.append([cn or "", pt or "", pn or "", "", "", "", "", "", True, "", "", ""])
            else:
                phone_raw = decrypt(phone_enc) if phone_enc else ""
                email = decrypt(email_enc) if email_enc else ""
                telegram_id = decrypt(tg_enc) if tg_enc else ""
                how_to_address = decrypt(how_enc) if how_enc else ""
                phone_display = _format_phone_display(phone_raw) or phone_raw or ""
                bv_display = BARRIER_VOTE_DISPLAY.get(bv, bv) if bv else ""
                vf_display = VOTE_FORMAT_DISPLAY.get(vf, vf) if vf else ""
                reg_ed_display = REGISTERED_ED_DISPLAY.get(reg_ed, reg_ed) if reg_ed else ""
                # Колонка «Ссылка ТГ» заполняется формулой ниже (пересчитывается при редактировании D/F)
                rows.append([
                    cn or "", pt or "", pn or "",
                    phone_display, email or "", telegram_id or "", "", how_to_address or "",
                    is_owner if is_owner is not None else True,
                    bv_display, vf_display, reg_ed_display,
                ])

    # Canary: вставить строку в блок помещения canary_row["premise_id"]
    if canary_row:
        pid = canary_row["premise_id"]
        insert_idx = -1
        for i, row in enumerate(rows):
            if row[0] == pid:
                insert_idx = i
        if insert_idx >= 0:
            cn, pt, pn = rows[insert_idx][0], rows[insert_idx][1], rows[insert_idx][2]
        else:
            with get_db() as db2:
                pr = db2.execute(text("SELECT cadastral_number, premises_type, premises_number FROM premises WHERE cadastral_number = :pid"), {"pid": pid}).fetchone()
                cn = pr[0] if pr else pid
                pt = pr[1] if pr else ""
                pn = pr[2] if pr else ""
        phone_display = _format_phone_display(canary_row["phone"]) or canary_row["phone"]
        canary_line = [
            cn or "", pt or "", pn or "",
            phone_display, "", canary_row["canary_telegram_id"], "", canary_row["how_to_address"],
            True, "", "", "",
        ]
        if insert_idx >= 0:
            rows.insert(insert_idx + 1, canary_line)
        else:
            rows.append(canary_line)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Контакты")
    ws.title = "Контакты"
    ws.append(CONTACTS_TEMPLATE_HEADERS_RU)
    for row in rows:
        ws.append(row)

    # SR-ADM08-006: колонка «Ссылка ТГ» — формула от D (телефон) и F (telegram_id), пересчитывается при редактировании
    g_col = COL_TG_LINK + 1
    for row_idx in range(2, len(rows) + 2):
        # F=telegram_id: если число — tg://user?id=; иначе https://t.me/username. D=телефон — https://t.me/+цифры
        formula = (
            f'=IF(F{row_idx}<>"",'
            f'IF(ISERROR(VALUE(F{row_idx})),'
            f'HYPERLINK("https://t.me/"&SUBSTITUTE(SUBSTITUTE(TRIM(F{row_idx}),"@","")," ",""),"\u2708"),'
            f'HYPERLINK("tg://user?id="&F{row_idx},"\u2708")),'
            f'IF(D{row_idx}<>"",'
            f'HYPERLINK("https://t.me/+"&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(D{row_idx}," ",""),"(",""),")",""),"-",""),"+",""),".",""),"\u2708"),""))'
        )
        ws.cell(row=row_idx, column=g_col).value = formula

    # Data Validation: выпадающие списки для позиции по шлагбаумам и формата голосования (openpyxl поддерживает)
    from openpyxl.worksheet.datavalidation import DataValidation
    max_row = len(rows) + 1
    dv_bv = DataValidation(
        type="list",
        formula1='"ЗА,Против,Не определился"',
        allow_blank=True,
        promptTitle="Позиция по шлагбаумам",
        prompt="Выберите: ЗА, Против, Не определился",
    )
    dv_bv.add(f"J2:J{max_row}")
    ws.add_data_validation(dv_bv)
    dv_vf = DataValidation(
        type="list",
        formula1='"Электронно,Бумага,Не определился"',
        allow_blank=True,
        promptTitle="Формат голосования",
        prompt="Выберите: Электронно, Бумага, Не определился",
    )
    dv_vf.add(f"K2:K{max_row}")
    ws.add_data_validation(dv_vf)
    dv_re = DataValidation(
        type="list",
        formula1='"Да,Нет"',
        allow_blank=True,
        promptTitle="Электронный дом",
        prompt="Зарегистрирован в Электронном доме: Да, Нет",
    )
    dv_re.add(f"L2:L{max_row}")
    ws.add_data_validation(dv_re)

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), len(rows)
